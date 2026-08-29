"""Le dossier de crédit Excel sur Enbridge : étalement, ratios, projection à formules, cotation.

Le classeur est celui qu'un analyste de crédit commercial prépare pour un comité : l'étalement
(les états financiers alignés sur dix exercices), les ratios de crédit, une projection à trois ans
dont les hypothèses sont des cellules modifiables (les formules recalculent), une grille de
cotation pondérée, et la feuille des clauses (covenants) proposées. La grille et ses seuils sont
des préceptes internes déclarés, pas une notation d'agence.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd

# la grille : (ratio, poids, seuils du score 5 au score 2 ; en dessous, score 1)
GRID = [
    ("dette_sur_ebitda", 0.30, [3.0, 4.0, 5.0, 6.0], "moins_est_mieux"),
    ("couverture_interets", 0.25, [6.0, 4.0, 3.0, 2.0], "plus_est_mieux"),
    ("dscr", 0.25, [2.0, 1.5, 1.2, 1.0], "plus_est_mieux"),
    ("ffo_sur_dette_pct", 0.20, [20.0, 15.0, 10.0, 6.0], "plus_est_mieux"),
]
LETTRES = [(4.5, "A"), (3.5, "BBB"), (2.5, "BB"), (1.5, "B"), (0.0, "CCC")]


def credit_ratios(h: pd.DataFrame) -> pd.DataFrame:
    """Les ratios de crédit annuels d'Enbridge, calculés depuis l'étalement SEC."""
    df = h.copy()
    df["ebitda"] = df["resultat_exploitation"] + df["amortissements"]
    df["dette_sur_ebitda"] = df["dette_long_terme"] / df["ebitda"]
    df["couverture_interets"] = df["ebitda"] / df["interets"]
    df["dscr"] = df["ebitda"] / (df["interets"] + df["remboursements_dette"])
    df["ffo_sur_dette_pct"] = 100.0 * df["flux_exploitation"] / df["dette_long_terme"]
    return df


def score_ratio(value: float, thresholds: list[float], direction: str) -> int:
    scores = [5, 4, 3, 2]
    for s, t in zip(scores, thresholds, strict=True):
        if (direction == "plus_est_mieux" and value >= t) or (direction == "moins_est_mieux" and value <= t):
            return s
    return 1


def scorecard(ratios_last: pd.Series) -> tuple[pd.DataFrame, float, str]:
    """La note pondérée du dernier exercice et sa lettre interne."""
    rows = []
    for name, weight, thresholds, direction in GRID:
        v = float(ratios_last[name])
        s = score_ratio(v, thresholds, direction)
        rows.append({"ratio": name, "valeur": v, "score_1_a_5": s, "poids": weight})
    table = pd.DataFrame(rows)
    note = float((table["score_1_a_5"] * table["poids"]).sum())
    lettre = next(letter for floor, letter in LETTRES if note >= floor)
    return table, note, lettre


def build_workbook(ratios: pd.DataFrame, table: pd.DataFrame, note: float, lettre: str,
                   dest: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "Lisez-moi"
    ws["A1"] = "Dossier de crédit : Enbridge Inc. (ENB), généré par clab"
    ws["A1"].font = bold
    ws["A3"] = "Étalement MESURÉ (SEC, companyfacts, M$ CAD) ; projection à FORMULES (hypothèses en jaune"
    ws["A4"] = "conceptuel : cellules B2:B4 de la feuille Projection) ; grille de cotation : préceptes internes"
    ws["A5"] = "déclarés, pas une notation d'agence. Le mémo bilingue est dans reports/."

    ws = wb.create_sheet("Etalement")
    ws.append(["exercice", *ratios.columns.tolist()])
    for c in ws[1]:
        c.font = bold
    for year, row in ratios.iterrows():
        ws.append([int(year), *[None if pd.isna(v) else round(float(v), 2) for v in row]])

    ws = wb.create_sheet("Projection")
    ws["A1"] = "Hypothèses (modifiables)"
    ws["A1"].font = bold
    last = ratios.index.max()
    base = ratios.loc[last]
    ws["A2"], ws["B2"] = "croissance des revenus", 0.03
    ws["A3"], ws["B3"] = "marge EBITDA (EBITDA/revenus)", round(float(base["ebitda"] / base["revenus"]), 4)
    ws["A4"], ws["B4"] = "taux d'intérêt effectif sur la dette", round(float(base["interets"] / base["dette_long_terme"]), 4)
    ws["A6"] = "Projection (M$ CAD)"
    ws["A6"].font = bold
    ws.append([])
    ws["A7"], ws["B7"], ws["C7"], ws["D7"] = "poste", f"{last} (réel)", f"{last + 1}", f"{last + 2}"
    for c in ws[7]:
        c.font = bold
    ws["A8"], ws["B8"] = "revenus", round(float(base["revenus"]), 1)
    ws["C8"], ws["D8"] = "=B8*(1+$B$2)", "=C8*(1+$B$2)"
    ws["A9"], ws["B9"] = "EBITDA", round(float(base["ebitda"]), 1)
    ws["C9"], ws["D9"] = "=C8*$B$3", "=D8*$B$3"
    ws["A10"], ws["B10"] = "dette long terme (stable, hypothèse)", round(float(base["dette_long_terme"]), 1)
    ws["C10"], ws["D10"] = "=B10", "=C10"
    ws["A11"], ws["B11"] = "intérêts", round(float(base["interets"]), 1)
    ws["C11"], ws["D11"] = "=C10*$B$4", "=D10*$B$4"
    ws["A12"], ws["B12"] = "dette/EBITDA", round(float(base["dette_sur_ebitda"]), 2)
    ws["C12"], ws["D12"] = "=C10/C9", "=D10/D9"
    ws["A13"], ws["B13"] = "couverture des intérêts", round(float(base["couverture_interets"]), 2)
    ws["C13"], ws["D13"] = "=C9/C11", "=D9/D11"

    ws = wb.create_sheet("Cotation")
    ws.append(["ratio", "valeur (dernier exercice)", "score 1-5", "poids"])
    for c in ws[1]:
        c.font = bold
    for _, r in table.iterrows():
        ws.append([r["ratio"], round(r["valeur"], 2), int(r["score_1_a_5"]), r["poids"]])
    ws.append([])
    ws.append(["note pondérée", round(note, 2), "lettre interne", lettre])
    ws[ws.max_row][0].font = bold

    ws = wb.create_sheet("Covenants")
    ws["A1"] = "Clauses proposées (illustratives, alignées sur la grille)"
    ws["A1"].font = bold
    for i, c in enumerate([
        "dette/EBITDA plafonnée à 6,0x, testée trimestriellement",
        "couverture des intérêts plancher à 2,5x",
        "information : états financiers trimestriels sous 45 jours, budget annuel sous 90 jours",
        "négatif : sûretés nouvelles limitées, distributions bloquées si dette/EBITDA > 5,5x",
    ], start=3):
        ws[f"A{i}"] = c
    for sheet in wb.worksheets:
        for col in ("A", "B", "C", "D"):
            sheet.column_dimensions[col].width = max(sheet.column_dimensions[col].width or 0, 34 if col == "A" else 14)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
